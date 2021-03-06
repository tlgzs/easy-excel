# !user/bin/env python3
# -*- coding: utf-8 -*-
# Author: ChaoqiYin
import xlrd
from openpyxl import load_workbook

from .import_sheet import ImportSheet
from ..import_pack.import_row import ImportRow
from .factory import type_factory


def check_is_xlsx(file_name):
    return file_name.lower().find('.xlsx') > 0


def turn_file_to_excel_workbook(file, sheet_no):
    is_file_path = False
    # 判断是路径还是file
    if isinstance(file, str):
        file_name = file
        is_file_path = True
    else:
        file_name = file.filename
    # 判断是不是xlsx
    if check_is_xlsx(file_name) is True:
        workbook = type_factory.get('xlsx')(load_workbook(file, data_only=True), sheet_no)
    else:
        factory = type_factory.get('xls')
        # 是路径的情况
        if is_file_path is True:
            workbook = factory(xlrd.open_workbook(file_name, formatting_info=True), sheet_no)
        else:
            workbook = factory(xlrd.open_workbook(file_contents=file.read(), formatting_info=True), sheet_no)
            try:
                file.close()
            except Exception as e:
                # 尝试关闭文件，不能关闭也无所谓
                pass
    return workbook


class ImportWorkbook(object):
    def __init__(self, file, converters):
        self.file = file
        self.converters = converters  # 转换类

    def add_converter(self, converter_key, func):
        '''
        设置转换类
        :param converter_key: 转换类型key
        :param func: 转换方法
        :return:
        '''
        if converter_key not in range(0, 7):
            raise Exception("converter_key must in the dict [0, 1, 2, 3, 4, 5, 6] keys")
        self.converters[converter_key] = func
        return self

    def do_import(self, parse_map, error_message_prefix='第{row_num}行', sheet_no=0, start_row_num=0, end_row_num=None,
                  row_del_class=None, row_validate_func=None, max_workers=None, title_row=None):
        '''
        导入启动方法
        :param parse_map: 解析的字典
        :param error_message_prefix: 报错提示的前缀文字, 默认是'第{row_num}'
        :param sheet_no: 解析的表格索引
        :param start_row_num: 从第几行开始解析
        :param end_row_num: 到第几行结束
        :param row_del_class: 默认的行处理类, 需要是ImportRow的子类
        :param row_validate_func: 行验证方法，接收4个参数：（行索引，行原始数据，行转换后的数据，parse_map），返回None或一个list，里面是该行的错误消息，会自动拼接上error_message_prefix
        :param max_workers: 异步线程数
        :param title_row: col_name根据行数去匹配index
        :return:
        '''
        rel_row_del_class = ImportRow if row_del_class is None else row_del_class
        excel = turn_file_to_excel_workbook(self.file, sheet_no)
        sheet = ImportSheet(self, excel, parse_map, error_message_prefix, sheet_no,
                            start_row_num, end_row_num, max_workers, rel_row_del_class, row_validate_func, title_row)
        return sheet.get_value()